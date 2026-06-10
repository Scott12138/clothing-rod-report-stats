"""
衣杆报表统计 v4
==================================================
运行方式：python3 remark_stats.py
  - 弹出文件选择窗口 → 选择 xlsx / csv 源表
  - 逐条解析"商家备注"，每条备注 → 结果表一行
  - 输出列（严格对齐目标表 10 列）：
      A:款式   B:颜色   C:米数总计   D:底座总数   E:全包围底座总数
      F:顶装半通总数   G:顶装全通总数   H:顶装转角总数
      I:中托总数   J:螺丝总数
      + 核对列：原始备注
  - 输出文件1：源文件名（明细表）.xlsx
  - 输出文件2：源文件名（米数统计表）.xlsx  ← 按款式+颜色聚合，格式与目标表一致
"""

import os, re, sys
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════
# 1. 文件选择
# ══════════════════════════════════════════════

def pick_file():
    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title="选择源表文件（xlsx / csv）",
        initialdir=os.path.expanduser("~/Desktop"),
        filetypes=[("Excel 文件", "*.xlsx"), ("CSV 文件", "*.csv"),
                   ("所有支持格式", "*.xlsx *.csv")],
    )
    root.destroy()
    return path

def show_error(msg):
    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    messagebox.showerror("错误", msg); root.destroy()

def show_info(msg):
    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    messagebox.showinfo("完成", msg); root.destroy()

SOURCE = pick_file()
if not SOURCE:
    print("未选择文件，退出。"); sys.exit(0)

ext = os.path.splitext(SOURCE)[1].lower()
if ext not in (".xlsx", ".csv"):
    show_error(f"不支持的格式：{ext}\n请选择 .xlsx 或 .csv 文件。"); sys.exit(1)

source_dir  = os.path.dirname(SOURCE)
source_base = os.path.splitext(os.path.basename(SOURCE))[0]
OUTPUT      = os.path.join(source_dir, f"{source_base}（明细表）.xlsx")
OUTPUT_STAT = os.path.join(source_dir, f"{source_base}（米数统计表）.xlsx")
print(f"📂 源文件：{SOURCE}")
print(f"📝 输出  ：{OUTPUT}")


# ══════════════════════════════════════════════
# 2. 读取源表
# ══════════════════════════════════════════════

df_src = pd.read_csv(SOURCE, encoding="utf-8-sig") if ext == ".csv" \
         else pd.read_excel(SOURCE)
df_src.columns = df_src.columns.str.strip()
if "商家备注" not in df_src.columns:
    show_error(f"源表中未找到[商家备注]列\n当前列：{list(df_src.columns)}"); sys.exit(1)

remarks = df_src["商家备注"].dropna()
remarks = remarks[remarks.astype(str).str.strip() != ""]
print(f"✅ 非空备注：{len(remarks)} 条")


# ══════════════════════════════════════════════
# 3. 解析逻辑
# ══════════════════════════════════════════════

# ── 款式关键词 ──
def extract_style(text):
    # 先做容错匹配（typo 变体），再统一映射回标准名
    for kw in ["X3150", "M335", "6号威法", "6号威发", "2mm"]:
        if kw in text:
            return kw
    # X315（漏了0）→ X3150
    if re.search(r"X315[^0]", text):
        return "X3150"
    return ""


# ── 颜色映射 ──
COLOR_RULES = [
    (r"瓷白|白色|白", "白"),
    (r"暮云灰|玄铁灰|灰色|灰", "灰"),
    (r"雅黑|路虎黑|黑色|黑", "黑"),
    (r"香槟|金色|金", "香槟/金"),
]
def extract_color(text):
    for pat, label in COLOR_RULES:
        if re.search(pat, text):
            return label
    return "其他"


# ── 辅助解析函数 ──

SKIP_RE = [r"^共\d+根$", r"^只要杆子", r"^纯杆子", r"^补发", r"^\d+$",
           r"^深度\d+$", r"^-\d+", r"^加点", r"^升级"]

def should_skip(tok):
    for pat in SKIP_RE:
        if re.match(pat, tok): return True
    if re.match(r"^[\u4e00-\u9fff]+$", tok): return True
    return False

def is_size_item(tok):
    return bool(re.match(r"^\d{2,4}\s*[*×xX]\s*\d+\s*根?$", tok))

def parse_qty(tok):
    """返回 (名称, 数量)，支持 底座*2 / 底座2个 / 螺丝135 三种格式"""
    # 格式1：名称*数量（可含可选单位）
    m = re.search(r"^(.+?)\s*[*×xX]\s*(\d+)\s*[个根]?$", tok)
    if m: return m.group(1).strip(), int(m.group(2))
    # 格式2：名称+数字+单位（如 底座12个、半通6个）
    m = re.search(r"^(.+?)(\d+)\s*[个根]?$", tok)
    if m: return m.group(1).strip(), int(m.group(2))
    return tok, 0

def split_tokens(text):
    text = text.replace("，", ",").replace("、", ",").replace("；", ",")
    return [t.strip() for t in re.split(r"[,\s]+", text) if t.strip()]

def normalize_size(tok):
    m = re.match(r"^(\d{2,4})\s*[*×xX]\s*(\d+)\s*根?$", tok)
    if m: return int(m.group(1)), int(m.group(2))
    return None, 0


def parse_remark(raw):
    raw = str(raw).strip()

    # ── 拆分款式与明细 ──────────────────────────
    if "：" in raw:
        spec_part, detail = raw.split("：", 1)
    elif ":" in raw:
        spec_part, detail = raw.split(":", 1)
    elif "，" in raw and is_size_item(raw.split("，", 1)[0]) is False:
        spec_part, detail = raw.split("，", 1)
    else:
        spec_part, detail = "", raw

    spec_part = spec_part.strip()

    # ── 折扣前缀（如 "-5："）→ 从 detail 再拆一次 ──
    if re.match(r"^[-\d]+$", spec_part):
        for sep in ["：", ":", "，"]:
            if sep in detail:
                spec_part, detail = detail.split(sep, 1)
                break
        spec_part = spec_part.strip()

    # ── 处理「补发」──
    # 纯补发单：整条备注以"补发"开头
    if spec_part.startswith("补发"):
        spec_part = spec_part.replace("补发", "").strip("：: ")
        # 如果 spec_part 被清空，从 detail 重新拆
        if not spec_part:
            for sep in ["：", ":", "，"]:
                if sep in detail:
                    spec_part, detail = detail.split(sep, 1)
                    break
            spec_part = spec_part.strip()
    else:
        # 嵌套补发段落在 detail 中 → 截去（避免重复统计）
        if "补发" in detail:
            detail = re.split(r"补发", detail)[0]

    # ── 识别款式、颜色 ──────────────────────────
    style = extract_style(spec_part)
    if not style:
        style = extract_style(raw)
    if style == "6号威发":
        style = "6号威法"
    color = extract_color(spec_part.replace(style, "") if style else spec_part)
    if color == "其他" and style:
        color = extract_color(raw.replace(style, "", 1) if style in raw else raw)

    tokens = split_tokens(detail)

    # ── 统计变量 ──
    total_mm   = 0.0    # 总毫米数（之后 /1000 → 米数）
    base_total = 0
    quanbao    = 0
    bantong    = 0
    quantong   = 0
    zhuanjiao  = 0
    zhongtuo   = 0
    luosi      = 0

    for tok in tokens:
        if should_skip(tok):
            continue

        # 尺寸项 → 累加米数
        if is_size_item(tok):
            length, qty = normalize_size(tok)
            total_mm += length * qty
            continue

        # ── 底座项（优先处理）──
        if "底座" in tok:
            name, qty = parse_qty(tok)
            base_total += qty
            if "全包围" in tok:
                quanbao += qty
            continue

        # ── 配件项（不含"底座"）──
        name, qty = parse_qty(tok)
        if qty == 0:
            continue   # 无法解析数量则跳过

        n = name
        if "螺丝" in n:
            luosi += qty
        else:
            # 顶装配件分类规则（互斥优先级：转角 > 全通 > 半通）
            # "半通转角顶装" 含 转角 → 只算顶装转角，不算顶装半通
            if "转角" in n:
                zhuanjiao += qty
            elif "全通" in n:
                quantong += qty
            elif "半通" in n:
                bantong += qty
            if "中托" in n:
                zhongtuo += qty

    return {
        "款式":            style,
        "颜色":            color,
        "米数总计":        round(total_mm / 1000, 2) if total_mm > 0 else "",
        "底座总数":        base_total if base_total > 0 else "",
        "全包围底座总数":   quanbao if quanbao > 0 else "",
        "顶装半通总数":     bantong if bantong > 0 else "",
        "顶装全通总数":     quantong if quantong > 0 else "",
        "顶装转角总数":     zhuanjiao if zhuanjiao > 0 else "",
        "中托总数":        zhongtuo if zhongtuo > 0 else "",
        "螺丝总数":        luosi if luosi > 0 else "",
        "_原始备注":        raw,
    }


# ══════════════════════════════════════════════
# 4. 批量解析
# ══════════════════════════════════════════════

OUT_COLS = [
    "款式", "颜色", "米数总计",
    "底座总数", "全包围底座总数",
    "顶装半通总数", "顶装全通总数", "顶装转角总数",
    "中托总数", "螺丝总数",
    "_原始备注",
]

rows = [parse_remark(r) for r in remarks]
df_out = pd.DataFrame(rows, columns=OUT_COLS)
print(f"✅ 解析完成，共 {len(df_out)} 行")


# ══════════════════════════════════════════════
# 5. 写 Excel
# ══════════════════════════════════════════════

DISPLAY_COLS = [
    "款式", "颜色", "米数总计",
    "底座总数", "全包围底座总数",
    "顶装半通总数", "顶装全通总数", "顶装转角总数",
    "中托总数", "螺丝总数",
    "原始备注（核对用）",
]

HDR_BG = "1F4E79"; SUB_BG = "2E75B6"
ODD_BG = "EBF3FB"; EVEN_BG = "FFFFFF"
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
RIGHT_ALIGN  = Alignment(horizontal="right",  vertical="center")
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

wb = Workbook()
ws = wb.active
ws.title = "明细表"
ws.freeze_panes = "A3"

ncols = len(DISPLAY_COLS)
col_letter = get_column_letter(ncols)

# 标题行
ws.merge_cells(f"A1:{col_letter}1")
tc = ws.cell(1, 1, f"{source_base} — 商家备注明细")
tc.font      = Font(bold=True, size=14, color="FFFFFF")
tc.fill      = PatternFill("solid", start_color=HDR_BG)
tc.alignment = CENTER_ALIGN
ws.row_dimensions[1].height = 30

# 表头行
for ci, col in enumerate(DISPLAY_COLS, 1):
    c = ws.cell(2, ci, col)
    c.font      = Font(bold=True, size=10, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=SUB_BG)
    c.alignment = CENTER_ALIGN
    c.border    = thin_border()
ws.row_dimensions[2].height = 26

# 数据行
NUM_COLS = {"米数总计", "底座总数", "全包围底座总数",
            "顶装半通总数", "顶装全通总数", "顶装转角总数",
            "中托总数", "螺丝总数"}
CENTER_COLS = {"款式", "颜色"}

for ri, row in df_out.iterrows():
    er = ri + 3
    bg = ODD_BG if ri % 2 == 0 else EVEN_BG
    for ci, key in enumerate(OUT_COLS, 1):
        val = row[key]
        if pd.isna(val): val = ""
        c = ws.cell(er, ci, val)
        c.font   = Font(size=10)
        c.fill   = PatternFill("solid", start_color=bg)
        c.border = thin_border()
        display_key = DISPLAY_COLS[ci - 1]
        if display_key in NUM_COLS:
            c.alignment = RIGHT_ALIGN
        elif display_key in CENTER_COLS:
            c.alignment = CENTER_ALIGN
        else:
            c.alignment = LEFT_ALIGN

# 列宽
COL_WIDTHS = {
    "款式": 10, "颜色": 8, "米数总计": 10,
    "底座总数": 10, "全包围底座总数": 14,
    "顶装半通总数": 12, "顶装全通总数": 12, "顶装转角总数": 12,
    "中托总数": 10, "螺丝总数": 10,
    "原始备注（核对用）": 55,
}
for ci, key in enumerate(DISPLAY_COLS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(key, 14)

wb.save(OUTPUT)
print(f"\n🎉 输出明细表：{OUTPUT}")


# ══════════════════════════════════════════════
# 6. 生成米数统计表（按款式+颜色聚合，格式与目标表一致）
# ══════════════════════════════════════════════

# 目标表固定结构：4 种款式，每种 4 色，款式间空行
STAT_STYLES = ["X3150", "M335", "6号威法", "2mm"]
STAT_COLORS = ["白", "灰", "黑", "香槟/金"]
STAT_COLS   = [
    "款式", "颜色", "米数总计",
    "底座总数", "全包围底座总数",
    "顶装半通总数", "顶装全通总数", "顶装转角总数",
    "中托总数", "螺丝总数",
]
NUM_STAT_COLS = STAT_COLS[2:]   # C~J 为数值列

# ── 聚合明细表 ──
AGG_KEYS = ["米数总计", "底座总数", "全包围底座总数",
            "顶装半通总数", "顶装全通总数", "顶装转角总数",
            "中托总数", "螺丝总数"]

# 把空字符串替换为 0，再做 groupby sum
df_agg = df_out[["款式", "颜色"] + AGG_KEYS].copy()
for k in AGG_KEYS:
    df_agg[k] = pd.to_numeric(df_agg[k], errors="coerce").fillna(0)

df_sum = df_agg.groupby(["款式", "颜色"], sort=False)[AGG_KEYS].sum().reset_index()

# ── 写统计表 Excel ──
wb2  = Workbook()
ws2  = wb2.active
ws2.title = "米数统计表"
ws2.freeze_panes = "A3"

ncols2      = len(STAT_COLS)
col_letter2 = get_column_letter(ncols2)

HDR_BG2 = "1F4E79"; SUB_BG2 = "2E75B6"
STYLE_BG  = "D6E4F0"   # 每个款式第一行（含款式名）的浅蓝背景
EMPTY_BG  = "F5F5F5"   # 空行背景
ODD_BG2   = "EBF3FB"; EVEN_BG2 = "FFFFFF"

# 大标题行
ws2.merge_cells(f"A1:{col_letter2}1")
tc2 = ws2.cell(1, 1, f"{source_base} — 按款式米数统计表")
tc2.font      = Font(bold=True, size=14, color="FFFFFF")
tc2.fill      = PatternFill("solid", start_color=HDR_BG2)
tc2.alignment = CENTER_ALIGN
ws2.row_dimensions[1].height = 30

# 表头行
for ci, col in enumerate(STAT_COLS, 1):
    c = ws2.cell(2, ci, col)
    c.font      = Font(bold=True, size=10, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=SUB_BG2)
    c.alignment = CENTER_ALIGN
    c.border    = thin_border()
ws2.row_dimensions[2].height = 26

# 数据行：按目标表结构填充（款式固定顺序，颜色固定顺序，中间空行）
data_row = 3   # 从第3行开始写数据
color_count = 0   # 当前款式已写颜色行数（用于斑马纹）

for si, style in enumerate(STAT_STYLES):
    color_count = 0
    style_has_data = False   # 标记该款式是否有任何有效行

    for color in STAT_COLORS:
        # 查聚合结果
        matched = df_sum[(df_sum["款式"] == style) & (df_sum["颜色"] == color)]

        # 米数为 0（无该款式+颜色的订单）→ 跳过，不写行
        mi_val = float(matched["米数总计"].iloc[0]) if len(matched) > 0 else 0
        if mi_val == 0:
            continue

        style_has_data = True

        # 斑马纹
        bg = ODD_BG2 if color_count % 2 == 0 else EVEN_BG2
        color_count += 1

        for ci, key in enumerate(STAT_COLS, 1):
            if key == "款式":
                val = style
            elif key == "颜色":
                val = color
            else:
                val = float(matched[key].iloc[0]) if len(matched) > 0 else 0
                # 米数保留2位小数，其余整数
                if key == "米数总计":
                    val = round(val, 2) if val != 0 else ""
                else:
                    val = int(val) if val != 0 else ""

            c = ws2.cell(data_row, ci, val)
            c.font   = Font(size=10)
            c.fill   = PatternFill("solid", start_color=bg)
            c.border = thin_border()
            if key in ("款式", "颜色"):
                c.alignment = CENTER_ALIGN
            elif key in NUM_STAT_COLS:
                c.alignment = RIGHT_ALIGN
            else:
                c.alignment = LEFT_ALIGN

        data_row += 1

    # 款式之间插入空行（该款式有数据 且 不是最后一个款式）
    if style_has_data and si < len(STAT_STYLES) - 1:
        # 检查后续款式是否还有数据，有才加空行
        has_next_data = any(
            float(df_sum[(df_sum["款式"] == ns) & (df_sum["颜色"] == nc)]["米数总计"].iloc[0])
            if len(df_sum[(df_sum["款式"] == ns) & (df_sum["颜色"] == nc)]) > 0 else 0
            for ns in STAT_STYLES[si+1:]
            for nc in STAT_COLORS
        )
        if has_next_data:
            for ci in range(1, ncols2 + 1):
                c = ws2.cell(data_row, ci, "")
                c.fill   = PatternFill("solid", start_color=EMPTY_BG)
                c.border = thin_border()
            ws2.row_dimensions[data_row].height = 8
            data_row += 1

# 列宽（与明细表一致）
COL_WIDTHS2 = {
    "款式": 10, "颜色": 8, "米数总计": 10,
    "底座总数": 10, "全包围底座总数": 14,
    "顶装半通总数": 12, "顶装全通总数": 12, "顶装转角总数": 12,
    "中托总数": 10, "螺丝总数": 10,
}
for ci, key in enumerate(STAT_COLS, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS2.get(key, 14)

wb2.save(OUTPUT_STAT)
print(f"🎉 输出米数统计表：{OUTPUT_STAT}")
show_info(
    f"统计完成！\n"
    f"共 {len(df_out)} 行明细\n\n"
    f"📋 明细表：{OUTPUT}\n"
    f"📊 米数统计表：{OUTPUT_STAT}"
)
